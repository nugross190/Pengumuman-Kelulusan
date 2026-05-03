"""
Excel parser for NILAI_SKL spreadsheet.

Reads the 'nilai jadi' sheet and produces a list of student records with
computed graduation status. Loaded once at app startup (the spreadsheet is
small enough — 406 students).

Layout assumptions (from the SMAN 5 Garut 2025/2026 SKL workbook):
- Sheet name: 'nilai jadi'
- Header rows occupy indices 0..7
- Data rows start at index 8
- Per-row layout:
    col 0: Nomor Urut
    col 1: NIS Induk
    col 2: NISN
    col 3: Nama
    cols 4..79: per-subject columns in groups of 4
              (Raport, PSAJ Tulis, PSAJ Praktek, Nilai Akhir)
- Lintas Minat (col 79) is INCLUDED in the average, but EXCLUDED from
  the KKM failure check (Lintas Minat scoring at SMAN 5 has a known
  issue where many students score below KKM through no fault of their
  own — so it should not block kelulusan).

The 'IDENTITAS' sheet supplies TTL and Nama Orang Tua, indexed by NISN.

If the spreadsheet structure changes, update SUBJECT_COLS below.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

log = logging.getLogger(__name__)

# Column index -> subject short code. These are the 'Nilai Akhir' columns
# for each subject. Lintas Minat (col 79) intentionally excluded.
SUBJECT_COLS: dict[int, str] = {
    7: "PAI",
    11: "PKn",
    15: "B. INDO",
    19: "MATEMATIKA",
    23: "SEJARAH INDONESIA",
    27: "B.INGGRIS",
    31: "SENI BUDAYA",
    35: "PJOK",
    39: "PRAKARYA",
    43: "B.SUNDA",
    47: "MATEMATIKA MINAT",
    51: "BIOLOGI",
    55: "FISIKA",
    59: "KIMIA",
    63: "SEJARAH MINAT",
    67: "GEOGRAFI",
    71: "SOSIOLOGI",
    75: "EKONOMI",
}

# Lintas Minat lives at col 79 ("Bahasa & Sastra Jepang/Arab/Inggris"). It
# counts toward the average but does NOT trigger Ditangguhkan if below KKM.
LINTAS_MINAT_COL = 79
LINTAS_MINAT_KEY = "LINTAS MINAT"

KKM = 75
HEADER_ROWS = 8
SHEET_NAME = "nilai jadi"
IDENTITAS_SHEET = "IDENTITAS"
IDENTITAS_HEADER_ROWS = 5  # data starts at row index 5


def _to_str(v) -> str:
    """Normalize a cell value to a clean string."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_float(v) -> Optional[float]:
    """Coerce a cell value to float, returning None if not numeric or zero/empty."""
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f == 0:
        return None
    return f


def _normalize_name(s: str) -> str:
    """Trim + collapse internal whitespace. Used for matching student names."""
    return " ".join(s.split())


def _parse_identitas(wb) -> dict[str, dict]:
    """Read the IDENTITAS sheet and return {nisn: {ttl, nama_ortu}}.

    Layout (from the SMAN 5 workbook):
      row 4: header (NO, NO INDUK, NISN, NO PESERTA, NAMA, ...,
             TEMPAT DAN TANGGAL LAHIR, NAMA ORANG TUA, ...)
      row 5+: data
      col 2 = NISN, col 7 = TTL, col 8 = Nama Orang Tua
    """
    ws = wb[IDENTITAS_SHEET]
    out: dict[str, dict] = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < IDENTITAS_HEADER_ROWS:
            continue
        if len(row) < 9:
            continue
        nisn = _to_str(row[2])
        if not nisn or nisn.lower() == "nan":
            continue
        out[nisn] = {
            "ttl": _to_str(row[7]),
            "nama_ortu": _to_str(row[8]),
        }
    return out


def parse_workbook(xlsx_path: Path) -> list[dict]:
    """
    Read the workbook and return a list of student records.

    Each record:
        {
          "nisn": "0079210473",
          "nis": "232410001",
          "nama": "ALDI RUSLAN",
          "jurusan": "MIPA" | "IPS" | "",
          "grades": { "PAI": 85.79, ... },
          "average": 86.22,
          "computed_status": "Lulus" | "Ditangguhkan",
          "failed_subjects": ["PKn", ...],   # only when Ditangguhkan
        }
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {xlsx_path}")

    log.info("Loading workbook: %s", xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in workbook. "
                         f"Available sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    identitas = _parse_identitas(wb) if IDENTITAS_SHEET in wb.sheetnames else {}

    students: list[dict] = []
    skipped_no_id = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < HEADER_ROWS:
            continue

        nomor = _to_str(row[0]) if len(row) > 0 else ""
        nis = _to_str(row[1]) if len(row) > 1 else ""
        nisn = _to_str(row[2]) if len(row) > 2 else ""
        nama = _to_str(row[3]) if len(row) > 3 else ""

        # Skip rows that aren't student records
        if not nomor or not nama or not nisn or nisn.lower() == "nan":
            skipped_no_id += 1
            continue

        nama = _normalize_name(nama)

        # Read all subject grades
        grades: dict[str, float] = {}
        failed: list[str] = []
        for col_idx, subj in SUBJECT_COLS.items():
            if col_idx >= len(row):
                continue
            val = _to_float(row[col_idx])
            if val is None:
                continue
            grades[subj] = round(val, 2)
            if val < KKM:
                failed.append(subj)

        # Lintas Minat: counted toward average, NOT toward failure check.
        if LINTAS_MINAT_COL < len(row):
            lm_val = _to_float(row[LINTAS_MINAT_COL])
            if lm_val is not None:
                grades[LINTAS_MINAT_KEY] = round(lm_val, 2)

        if not grades:
            # Student row with no graded subjects — skip
            continue

        # Determine jurusan from filled peminatan slots
        has_mipa = any(s in grades for s in ("BIOLOGI", "FISIKA", "KIMIA"))
        has_ips = any(s in grades for s in ("GEOGRAFI", "SOSIOLOGI", "EKONOMI"))
        if has_mipa:
            jurusan = "MIPA"
        elif has_ips:
            jurusan = "IPS"
        else:
            jurusan = ""

        average = round(sum(grades.values()) / len(grades), 2) if grades else 0.0
        status = "Ditangguhkan" if failed else "Lulus"

        ident = identitas.get(nisn) or identitas.get(nisn.zfill(10)) or {}

        students.append({
            "nisn": nisn,
            "nis": nis,
            "nama": nama,
            "jurusan": jurusan,
            "grades": grades,
            "average": average,
            "computed_status": status,
            "failed_subjects": failed,
            "ttl": ident.get("ttl", ""),
            "nama_ortu": ident.get("nama_ortu", ""),
        })

    wb.close()
    log.info(
        "Parsed %d students (skipped %d non-record rows)",
        len(students), skipped_no_id,
    )
    return students


def build_index(students: list[dict]) -> dict[str, dict]:
    """Index students by NISN for O(1) lookup."""
    return {s["nisn"]: s for s in students}
